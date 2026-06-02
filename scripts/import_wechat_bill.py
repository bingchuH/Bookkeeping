#!/usr/bin/env python3
import argparse
import collections
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

from bookkeeping import valid_categories

STANDARD_FIELDS = [
    "transaction_time",
    "type",
    "category_level1",
    "category_level2",
    "category",
    "amount",
    "currency",
    "transaction_object",
    "account",
    "target_account",
    "participant",
    "note",
]

WECHAT_FIELD_MAP = {
    "交易时间": "transaction_time",
    "交易类型": "type_hint",
    "交易对方": "transaction_object",
    "商品": "note",
    "收/支": "direction",
    "金额(元)": "amount",
    "支付方式": "account",
    "当前状态": "status",
}


def clean_text(value):
    text = str(value or "").strip()
    return "" if text == "/" else text


def split_category(category):
    if not category:
        return "", ""
    parts = category.split("/", 1)
    return parts[0], parts[1] if len(parts) > 1 else ""


def normalize_payment_method(value, direction="", status=""):
    payment = clean_text(value)
    if payment:
        return payment
    if clean_text(direction) == "收入" and clean_text(status) == "已存入零钱":
        return "零钱"
    return payment


def account_name(row, account_map):
    key = normalize_payment_method(row.get("支付方式"), row.get("收/支"), row.get("当前状态"))
    mapped = account_map.get(key)
    if mapped:
        return mapped
    return re.sub(r"\s+", "", key)


def transaction_object(row):
    return clean_text(row.get("交易对方"))


def note_text(row):
    product = clean_text(row.get("商品"))
    tx_type = clean_text(row.get("交易类型"))
    status = clean_text(row.get("当前状态"))
    parts = []
    if product:
        parts.append(product)
    if tx_type:
        parts.append(f"微信交易类型:{tx_type}")
    if status:
        parts.append(f"状态:{status}")
    return "；".join(parts)


def builtin_classification(row):
    source = clean_text(row.get("交易类型"))
    direction = clean_text(row.get("收/支"))

    # Keep classifications explicitly encoded by WeChat's structural fields.
    # Merchant and product keyword classification belongs to the standard CSV flow.
    if "退款" in source:
        return "收入", "退款"
    if source == "二手交易款":
        return "收入", "二手闲置"
    if source.startswith("微信红包") and direction == "收入":
        return "收入", "红包"
    if source.startswith("微信红包") and direction == "支出":
        return "支出", "人情往来/礼金"

    if direction == "收入":
        return "收入", ""
    if direction != "支出":
        return "", ""

    return "支出", ""


def classify(row, account_map):
    tx_type, category = builtin_classification(row)
    if not tx_type:
        return None, f"未映射的不计收支：{clean_text(row.get('交易类型')) or '(空)'}"

    if category and category not in valid_categories():
        category = ""
    level1, level2 = split_category(category)
    tx_time = row.get("交易时间")
    if hasattr(tx_time, "strftime"):
        tx_time = tx_time.strftime("%Y-%m-%d %H:%M:%S")
    return {
        "transaction_time": clean_text(tx_time),
        "type": tx_type,
        "category_level1": level1,
        "category_level2": level2,
        "category": category,
        "amount": row.get("金额(元)"),
        "currency": "CNY",
        "transaction_object": transaction_object(row),
        "account": account_name(row, account_map),
        "target_account": "",
        "participant": "自己",
        "note": note_text(row),
    }, None


def read_wechat_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_index = None
    header = None
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [clean_text(value) for value in row]
        if values and values[0] == "交易时间" and "金额(元)" in values:
            header_index = index
            header = values
            break
    if header_index is None:
        raise SystemExit("未找到微信支付明细表头：交易时间")

    rows = []
    for raw in sheet.iter_rows(min_row=header_index + 1, values_only=True):
        if not any(value is not None for value in raw):
            continue
        row = dict(zip(header, raw))
        if not row.get("交易时间"):
            continue
        rows.append(row)
    return header_index, header, rows


def classify_rows(rows, account_map):
    results = []
    skipped = collections.Counter()
    for row in rows:
        record, reason = classify(row, account_map)
        if record:
            results.append((row, record))
        else:
            skipped[reason] += 1
    return results, skipped


def analyze(rows):
    print(f"原始记录: {len(rows)}")
    for field in ["交易类型", "收/支", "支付方式", "当前状态", "交易对方"]:
        counter = collections.Counter(clean_text(row.get(field)) for row in rows)
        print(f"\n{field} 分布:")
        for name, count in counter.most_common(20):
            print(f"  {count:>3} {name or '(空)'}")
    payment_methods = collections.Counter(
        normalize_payment_method(row.get("支付方式"), row.get("收/支"), row.get("当前状态"))
        for row in rows
    )
    print("\n账户映射模板:")
    print(json.dumps({name: "" for name in payment_methods}, ensure_ascii=False, indent=2))
    print("\n内置字段映射:")
    print(json.dumps(WECHAT_FIELD_MAP, ensure_ascii=False, indent=2))


def write_standard_csv(rows, output):
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=STANDARD_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    return output


def read_json_mapping(value):
    if not value:
        return {}
    path = Path(value)
    text = path.read_text(encoding="utf-8") if path.exists() else value
    data = json.loads(text)
    if not isinstance(data, dict):
        raise SystemExit("mapping must be a JSON object")
    return {str(key): str(item) for key, item in data.items()}


def main():
    parser = argparse.ArgumentParser(description="Analyze and clean WeChat Pay XLSX bills to standard CSV")
    parser.add_argument("path")
    parser.add_argument("--output", default="/tmp/wechat_standard.csv")
    parser.add_argument("--account-map", help="JSON object or JSON file mapping WeChat payment methods to local account names")
    parser.add_argument("--analyze-only", action="store_true")
    args = parser.parse_args()

    account_map = read_json_mapping(args.account_map)
    header_line, header, rows = read_wechat_xlsx(args.path)
    print(f"表头行: {header_line}")
    print("表头:", ", ".join(header) if header else "(空)")
    analyze(rows)
    if args.analyze_only:
        return

    classified, skipped = classify_rows(rows, account_map)
    cleaned = [record for _, record in classified]
    output = write_standard_csv(cleaned, args.output)
    print(f"\n标准 CSV: {output}")
    print(f"清洗后记录: {len(cleaned)}")
    print("跳过记录:")
    for reason, count in skipped.most_common():
        print(f"  {count:>3} {reason}")
    print("清洗后类型分布:", dict(collections.Counter(row["type"] for row in cleaned)))
    print("清洗后账户分布:", dict(collections.Counter(row["account"] for row in cleaned)))
    print("空分类:", sum(1 for row in cleaned if not row["category"]))


if __name__ == "__main__":
    main()
